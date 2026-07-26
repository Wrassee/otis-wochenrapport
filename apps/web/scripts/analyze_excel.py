from openpyxl import load_workbook
import sys

# Force UTF-8 for stdout
sys.stdout.reconfigure(encoding='utf-8')

files = [
    'D:/DOKUMENTUMOK/OTIS/RAPPORT/Wochen rapport stu 2026 KW28.xlsx',
    'D:/DOKUMENTUMOK/OTIS/RAPPORT/Wochen rapport stu 2026 KW29.xlsx'
]

for f in files:
    print(f"\n{'='*80}")
    print(f"FILE: {f.split('/')[-1]}")
    print('='*80)
    
    wb = load_workbook(f, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        max_row = min(ws.max_row, 55)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=False), 1):
            vals = []
            for c in row:
                v = c.value
                if v is not None:
                    cell_ref = f"{c.column_letter}{row_idx}"
                    sv = str(v).replace('\n', ' | ')
                    if len(sv) > 80:
                        sv = sv[:80] + '...'
                    vals.append(f"{cell_ref}={sv}")
            if vals:
                print(f"  R{row_idx}: {' | '.join(vals)}")
    
    print()
