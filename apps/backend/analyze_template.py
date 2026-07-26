"""Analyze the Excel template structure - output to file."""
import sys
from openpyxl import load_workbook

wb = load_workbook('templates/template.xlsx')

with open('template_analysis.txt', 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n{'='*60}\n")
        f.write(f"=== {sheet_name} ===\n")
        f.write(f"Dimensions: {ws.dimensions}\n")
        f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n\n")
        
        for r in range(1, min(ws.max_row + 1, 51)):
            cells = []
            for c in range(1, min(ws.max_column + 1, 26)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    s = str(v).replace('\n', ' | ')[:80]
                    cl = chr(64 + c) if c <= 26 else f"C{c}"
                    cells.append(f"{cl}={s}")
            if cells:
                f.write(f"  R{r}: {' | '.join(cells)}\n")
        
        f.write("\nMerged cells:\n")
        for mc in sorted(ws.merged_cells.ranges, key=str):
            f.write(f"  {mc}\n")
    
    f.write("\nDone.\n")
