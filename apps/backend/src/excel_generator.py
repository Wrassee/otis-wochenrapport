"""
Excel Generator: Fills in the OTIS Wochenrapport template
with time entries and expense data.
"""

import os
from datetime import datetime, timedelta
from typing import Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Path to the template
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "templates",
    "template.xlsx",
)

# Weekday column mapping for Spesenrapport (D=Monday, E=Tuesday, etc.)
SPESEN_DAY_COLUMNS = {
    0: "D",  # Monday (LU)
    1: "E",  # Tuesday (MA)
    2: "F",  # Wednesday (ME)
    3: "G",  # Thursday (JE)
    4: "H",  # Friday (VE)
    5: "I",  # Saturday (SA)
    6: "J",  # Sunday (DI)
}

# Column mapping for Stundenrapport activity codes
ACTIVITY_COLUMNS = {
    "NK": "J",
    "S": "J",
    "T": "J",
    "T Clot": "K",
    "O": "L",
    "QI": "M",
    "I04": "N",
    "I5S": "N",
    "I5Q": "N",
    "I5T": "N",
    "I5A": "N",
    "A01": "N",
    "A02": "N",
    "A03": "N",
    "A04": "N",
    "A05": "N",
    "A07": "N",
    "VM": "O",
    "VP": "P",
    "NM": "Q",
    "NTC": "Q",
    "NF": "Q",
    "VC": "Q",
    "QI SCOTT": "R",
}

# Expense type → Spesenrapport row number mapping
# Column C contains the type identifier (2, 3, 4, 5, 6, 7, 28)
# We write "1" in the day column (D-J) for applicable expenses
# For Privatfahrzeug (row 33), we write the km count instead of "1"
EXPENSE_ROWS = {
    "entschaedigung_10h": 26,  # Entschädigung >= 10h (C26=2)
    "hotel": 27,               # Hotel (C27=3)
    "transport": 28,           # Transport (3) (C28=4)
    "pikettdienst": 29,        # Pikettdienst (C29=5)
    "entschaedigung_pikett": 30,  # Entschädigung Pikett (C30=6)
    "material": 31,            # Material (C31=7)
    "privatfahrzeug": 33,      # Privatfahrzeug (C33=28)
}


def _get_monday_of_week(year: int, week_number: int) -> datetime:
    """Get the Monday datetime for a given ISO week."""
    jan4 = datetime(year, 1, 4)
    day_offset = jan4.weekday()  # Monday=0
    monday = jan4 - timedelta(days=day_offset) + timedelta(weeks=week_number - 1)
    return monday


def _get_excel_day_column(weekday: int) -> str:
    """Get the Excel column letter for a given weekday in Spesenrapport."""
    return SPESEN_DAY_COLUMNS.get(weekday, "D")


def generate_excel(
    year: int,
    week_number: int,
    personnel_number: str,
    full_name: str,
    entries: list[dict],
    expenses: Optional[list[dict]] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Generate a filled Excel file from the template.

    Args:
        year: The year (e.g., 2026)
        week_number: The ISO week number (e.g., 29)
        personnel_number: The employee's personnel number
        full_name: The employee's full name
        entries: List of time entry dictionaries with keys:
            - date: str (YYYY-MM-DD)
            - start_time: float (decimal hours)
            - duration: float (decimal hours)
            - anlagenummer: str (optional)
            - project_id: str (optional)
            - address: str (optional)
            - activity_code: str (optional)
            - is_lunch: bool (optional)
            - zone: int (optional)
        expenses: Optional list of expense dictionaries with keys:
            - date: str (YYYY-MM-DD)
            - expense_type: str (see EXPENSE_ROWS keys)
            - value: number (1 for most, km count for privatfahrzeug)
        output_path: Optional path to save the file

    Returns:
        The Excel file content as bytes
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    # Copy template
    wb = load_workbook(TEMPLATE_PATH)

    # ========================
    # STUNDENRAPPORT (Hours Report)
    # ========================
    ws = wb["Stundenrapport"]

    # Fill header (Row 2-3)
    # C2 = Personnel Number (top-left of C2, standalone cell)
    ws["C2"] = personnel_number
    # E2 = Last Name (E2:F2 is merged, E2 is top-left → valid)
    # H2 = First Name (H2:J2 is merged, H2 is top-left → valid)
    name_parts = full_name.strip().split(" ", 1)
    if len(name_parts) == 2:
        ws["E2"] = name_parts[0]  # Last name
        ws["H2"] = name_parts[1]  # First name
    else:
        ws["E2"] = full_name
    # L2 = Month, N2 = Year
    week_monday = _get_monday_of_week(year, week_number)
    ws["L2"] = week_monday.month
    ws["N2"] = year

    # Set L3 = week reference
    ws["L3"] = week_number

    # ========================
    # SECOND BLOCK HEADER (rows 28-29) — same info as first block
    # This is a second copy of the form on the same sheet for printing
    # ========================
    # C28 = Personnel Number (B28 has the label "No. Pers. / Nr. Pers. / No. Pers.")
    ws["C28"] = personnel_number
    # E28 = Last Name (D28 has label, E28:F28 is the value area — not merged)
    # H28 = First Name (G28 has label, H28:J28 is merged for the value)
    if len(name_parts) == 2:
        ws["E28"] = name_parts[0]  # Last name
        ws["H28"] = name_parts[1]  # First name
    else:
        ws["E28"] = full_name
    # L28 = Month (K28 has the label "Mois:/Monat:/Mese:")
    ws["L28"] = week_monday.month
    # N28 = Year (M28 has the label "Année:/Jahr:/Anno:")
    ws["N28"] = year
    # L29 = week reference (same as L3)
    ws["L29"] = week_number

    # Filter non-lunch entries and sort by date and time
    work_entries = [
        e for e in entries
        if not e.get("is_lunch", False)
    ]
    work_entries.sort(key=lambda e: (e.get("date", ""), e.get("start_time", 0)))

    # Fill entries starting from row 8
    # IMPORTANT: Only fill rows 8-22 (max 15 entries) to avoid overwriting
    # the legend section at rows 24-25
    MAX_ENTRY_ROWS = 15
    START_ROW = 8
    current_row = START_ROW
    entry_count = 0

    for entry in work_entries:
        if current_row >= START_ROW + MAX_ENTRY_ROWS:
            break

        date_str = entry.get("date", "")
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            day_of_month = dt.day
        except (ValueError, TypeError):
            day_of_month = ""

        # Column A: Day of month (Jour/Tag/Giorno)
        ws.cell(row=current_row, column=1, value=day_of_month)

        # Column B: Anlagenummer (No. Appareil)
        anlagenummer = entry.get("anlagenummer", "")
        if anlagenummer:
            ws.cell(row=current_row, column=2, value=anlagenummer)

        # Column D: Project ID (No. Projet)
        project_id = entry.get("project_id", "")
        if project_id:
            ws.cell(row=current_row, column=4, value=project_id)

        # Column F: Address
        address = entry.get("address", "")
        if address:
            ws.cell(row=current_row, column=6, value=address)

        # Column H: Start time (Début/Anfang)
        start_time = entry.get("start_time", 0)
        ws.cell(row=current_row, column=8, value=start_time)

        # Column I: Duration (Durée/Dauer)
        duration = entry.get("duration", 0)
        ws.cell(row=current_row, column=9, value=duration)

        # Activity code column (J-R)
        activity_code = entry.get("activity_code", "")
        if activity_code and activity_code in ACTIVITY_COLUMNS:
            col_letter = ACTIVITY_COLUMNS[activity_code]
            col_index = ord(col_letter) - ord("A") + 1  # Convert to 1-based index
            ws.cell(row=current_row, column=col_index, value="ü")

        current_row += 1
        entry_count += 1

    # ========================
    # SPESENRAPPORT (Expense Report)
    # ========================
    ws2 = wb["Spesenrapport"]

    # Fill header
    ws2["B7"] = personnel_number
    ws2["G7"] = full_name  # G7:J7 is merged, G7 is top-left → valid

    # Date range
    monday = _get_monday_of_week(year, week_number)
    friday = monday + timedelta(days=4)
    ws2["E5"] = monday.strftime("%d.%m.%Y")   # E5:F5 merged, E5 is top-left
    ws2["I5"] = friday.strftime("%d.%m.%Y")   # I5:J5 merged, I5 is top-left

    # Calculate the highest zone for each day
    day_zones: dict[int, int] = {}  # weekday -> highest zone
    for entry in entries:
        date_str = entry.get("date", "")
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            weekday = dt.weekday()
        except (ValueError, TypeError):
            continue

        zone = entry.get("zone", 0)
        if zone and zone > day_zones.get(weekday, 0):
            day_zones[weekday] = zone

    # Fill zone marks in Spesenrapport
    # Zone rows (each row has the zone code in column C):
    # Row 10: Zone 1 (Z1)
    # Row 12: Zone 2a (Z2)
    # Row 15: Zone 3a (Z3)
    # Row 18: Zone 4a (Z4)
    zone_rows = {1: 10, 2: 12, 3: 15, 4: 18}

    for weekday, zone in day_zones.items():
        if zone in zone_rows:
            row = zone_rows[zone]
            col_letter = _get_excel_day_column(weekday)
            col_index = ord(col_letter) - ord("A") + 1
            ws2.cell(row=row, column=col_index, value=1)

    # ========================
    # EXPENSES (Spesenrapport)
    # ========================
    if expenses:
        # Group expenses by weekday for efficient writing
        # expenses dict: {weekday: {expense_type: value}}
        expense_by_day: dict[int, dict[str, float]] = {}
        for exp in expenses:
            date_str = exp.get("date", "")
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                weekday = dt.weekday()
            except (ValueError, TypeError):
                continue

            if weekday not in expense_by_day:
                expense_by_day[weekday] = {}

            exp_type = exp.get("expense_type", "")
            exp_value = exp.get("value", 1)
            expense_by_day[weekday][exp_type] = float(exp_value)

        # Write expense values to the appropriate rows and day columns
        for weekday, day_expenses in expense_by_day.items():
            col_letter = _get_excel_day_column(weekday)
            col_index = ord(col_letter) - ord("A") + 1

            for exp_type, value in day_expenses.items():
                if exp_type in EXPENSE_ROWS:
                    row = EXPENSE_ROWS[exp_type]
                    ws2.cell(row=row, column=col_index, value=value)

    # ========================
    # FOOTER: Datum + Unterschrift (Spesenrapport bottom)
    # ========================
    # E36 = Date value (C36 has the label "Date / Datum / Data :")
    today = datetime.now()
    ws2["E36"] = today.strftime("%d.%m.%Y")

    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    content = output.read()

    # Optionally save to file
    if output_path:
        with open(output_path, "wb") as f:
            f.write(content)

    return content
